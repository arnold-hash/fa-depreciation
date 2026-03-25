"""
build_fa_schedule.py
Retro Labs Ltd — FA Schedule Rebuild Module

Called by fa_depreciation.py step 5, or standalone:
  python build_fa_schedule.py

Pulls live data from Snowflake (FIVETRAN.XERO.ASSET + XERO.JOURNAL),
rebuilds FA_Schedule_FINAL.xlsx (4 tabs), uploads to Google Drive,
returns the shareable file URL.
"""

import os, io, sys, datetime, calendar
import snowflake.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from google.oauth2 import service_account
from googleapiclient.discovery import build as gdrive_build
from googleapiclient.http import MediaIoBaseUpload
import json

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
GREEN      = "E2EFDA"
AMBER      = "FFF2CC"
RED_LIGHT  = "FCE4D6"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"

TODAY = datetime.date.today()

# ─── SNOWFLAKE DATA FETCH ─────────────────────────────────────────────────────

def fetch_assets(conn):
    cur = conn.cursor()
    cur.execute("""
        SELECT
            a.ASSET_NUMBER,
            a.ASSET_NAME,
            at.ASSET_TYPE_NAME,
            TO_DATE(a.PURCHASE_DATE)          AS PURCHASE_DATE,
            a.PURCHASE_PRICE                  AS COST,
            a.ASSET_STATUS,
            a.DEPRECIATION_METHOD,
            a.DEPRECIATION_EFFECTIVE_LIFE_YEARS,
            a.ACCOUNTING_BOOK_VALUE,
            TO_DATE(a.DISPOSAL_DATE)          AS DISPOSAL_DATE
        FROM FIVETRAN.XERO.ASSET a
        LEFT JOIN FIVETRAN.XERO.ASSET_TYPE at ON a.ASSET_TYPE_ID = at.ASSET_TYPE_ID
        ORDER BY a.PURCHASE_DATE, a.ASSET_NUMBER
    """)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def fetch_dep_journals(conn):
    """
    Returns monthly dep totals: {(year, month): {"CE": float, "OE": float, "total": float}}
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT
            YEAR(j.JOURNAL_DATE)  AS YR,
            MONTH(j.JOURNAL_DATE) AS MO,
            jl.ACCOUNT_CODE,
            SUM(jl.NET_AMOUNT)    AS AMOUNT
        FROM FIVETRAN.XERO.JOURNAL j
        JOIN FIVETRAN.XERO.JOURNAL_LINE jl ON j.JOURNAL_ID = jl.JOURNAL_ID
        WHERE jl.ACCOUNT_CODE IN ('721','711')
          AND jl.NET_AMOUNT < 0
        GROUP BY 1,2,3
        ORDER BY 1,2
    """)
    result = {}
    for (yr, mo, acct, amt) in cur.fetchall():
        key = (int(yr), int(mo))
        if key not in result:
            result[key] = {"CE": 0.0, "OE": 0.0, "total": 0.0}
        val = abs(float(amt))
        if acct == "721":
            result[key]["CE"] += val
        else:
            result[key]["OE"] += val
        result[key]["total"] += val
    return result


# ─── DEP CALCULATION ─────────────────────────────────────────────────────────

def dep5(pdate, cost, yr, mo, disposal_date=None):
    if not pdate or not cost:
        return 0.0
    e = (yr - pdate.year) * 12 + (mo - pdate.month) + 1
    if e <= 0 or e > 60:
        return 0.0
    if disposal_date and (yr, mo) > (disposal_date.year, disposal_date.month):
        return 0.0
    return round(float(cost) / 60, 2)


# ─── EXCEL BUILD ─────────────────────────────────────────────────────────────

def hdr(ws, r, c, val, bg=DARK_BLUE, fg=WHITE, bold=True, size=9, wrap=False):
    cell = ws.cell(r, c, val)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return cell

def dat(ws, r, c, val, bg=WHITE, bold=False, size=9, fmt=None, center=False):
    cell = ws.cell(r, c, val)
    cell.font      = Font(name="Arial", bold=bold, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    if fmt:
        cell.number_format = fmt
    return cell


def build_workbook(assets, dep_journals):
    wb = Workbook()

    # ── All months from earliest asset to Jul 2026 ──
    earliest = min((a["PURCHASE_DATE"] for a in assets if a["PURCHASE_DATE"]), default=datetime.date(2021,10,1))
    all_months = []
    d = datetime.date(earliest.year, earliest.month, 1)
    end_month = datetime.date(2026, 7, 1)
    while d <= end_month:
        all_months.append((d.year, d.month))
        d += relativedelta_simple(d)

    active   = [a for a in assets if a["ASSET_STATUS"] == "Registered"]
    disposed = [a for a in assets if a["ASSET_STATUS"] == "Disposed"]

    # ────────────────────────────────────────────────
    # TAB 1: SUMMARY
    # ────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col, w in [("B",28),("C",16),("D",16),("E",16),("F",16),("G",16)]:
        ws.column_dimensions[col].width = w

    ws.merge_cells("B1:G1")
    c = ws["B1"]
    c.value = "RETRO LABS LTD — FIXED ASSET DEPRECIATION SCHEDULE"
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = f"Auto-generated {TODAY.strftime('%d %B %Y')} by GitHub Actions  |  5yr Straight-line  |  FY 1 Aug – 31 Jul"
    c.font = Font(name="Arial", size=9, color=WHITE, italic=True)
    c.fill = PatternFill("solid", fgColor=MID_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 8

    # KPI row
    total_cost = sum(float(a["COST"] or 0) for a in active)
    last_period = max(dep_journals.keys()) if dep_journals else (TODAY.year, TODAY.month)
    last_total  = dep_journals.get(last_period, {}).get("total", 0.0)
    kpis = [
        ("REGISTERED ASSETS", f"{len(active)} assets"),
        ("TOTAL COST BASIS",  f"£{total_cost:,.2f}"),
        ("LAST POSTED TOTAL", f"£{last_total:,.2f}"),
        ("LAST PERIOD",       datetime.date(last_period[0], last_period[1], 1).strftime("%b %Y")),
        ("SCHEDULE UPDATED",  TODAY.strftime("%d %b %Y")),
        ("METHOD",            "5yr SL / Full month"),
    ]
    for i, (title, val) in enumerate(kpis):
        col = i + 2
        ws.cell(4, col).value = title
        ws.cell(4, col).font  = Font(name="Arial", bold=True, size=8, color=WHITE)
        ws.cell(4, col).fill  = PatternFill("solid", fgColor=MID_BLUE)
        ws.cell(4, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(5, col).value = val
        ws.cell(5, col).font  = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
        ws.cell(5, col).fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(5, col).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 8

    # Period table
    hdr(ws, 7, 2, "PERIOD",       MID_BLUE, WHITE, size=9)
    hdr(ws, 7, 3, "CE (721) £",   MID_BLUE, WHITE, size=9)
    hdr(ws, 7, 4, "OE (711) £",   MID_BLUE, WHITE, size=9)
    hdr(ws, 7, 5, "TOTAL DEP £",  MID_BLUE, WHITE, size=9)
    hdr(ws, 7, 6, "SOURCE",       MID_BLUE, WHITE, size=9)
    hdr(ws, 7, 7, "STATUS",       MID_BLUE, WHITE, size=9)
    ws.row_dimensions[7].height = 16

    recent_periods = sorted(dep_journals.keys(), reverse=True)[:6]
    for i, key in enumerate(reversed(recent_periods)):
        r = 8 + i
        d = dep_journals[key]
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        period_str = datetime.date(key[0], key[1], 1).strftime("%b %Y")
        dat(ws, r, 2, period_str,   bg=bg, bold=True, size=9, center=True)
        dat(ws, r, 3, d["CE"],      bg=bg, size=9, center=True, fmt="£#,##0.00")
        dat(ws, r, 4, d["OE"],      bg=bg, size=9, center=True, fmt="£#,##0.00")
        dat(ws, r, 5, d["total"],   bg=bg, bold=True, size=9, center=True, fmt="£#,##0.00")
        dat(ws, r, 6, "Xero FA auto", bg=bg, size=9, center=True)
        dat(ws, r, 7, "Posted", bg=GREEN, bold=True, size=9, center=True)
        ws.row_dimensions[r].height = 15

    # ────────────────────────────────────────────────
    # TAB 2: ASSET REGISTER
    # ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Asset Register")
    ws2.sheet_view.showGridLines = False
    for col, (title, width) in enumerate([
        ("FA #",12),("Asset Name",30),("Type",10),("Purchase Date",14),
        ("Cost £",11),("Monthly dep £",13),("Book Value £",13),
        ("End of Life",12),("Status",11)
    ], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
        hdr(ws2, 1, col, title, DARK_BLUE, WHITE, size=9, wrap=True)
    ws2.row_dimensions[1].height = 20

    for i, a in enumerate(assets):
        r  = i + 2
        pd = a["PURCHASE_DATE"]
        cost = float(a["COST"] or 0)
        monthly = dep5(pd, cost, TODAY.year, TODAY.month, a.get("DISPOSAL_DATE"))
        eol = datetime.date(pd.year + 5, pd.month, pd.day) if pd else None

        if a["ASSET_STATUS"] == "Disposed": bg = RED_LIGHT
        elif i % 2 == 0: bg = LIGHT_BLUE
        else: bg = WHITE

        vals = [
            a["ASSET_NUMBER"], a["ASSET_NAME"],
            "CE" if "Computer" in (a.get("ASSET_TYPE_NAME") or "") else "OE",
            pd.isoformat() if pd else "",
            cost, monthly,
            float(a["ACCOUNTING_BOOK_VALUE"] or 0),
            eol.strftime("%b %Y") if eol and a["ASSET_STATUS"] != "Disposed" else "Disposed",
            a["ASSET_STATUS"],
        ]
        fmts = [None,None,None,None,"£#,##0.00","£#,##0.00","£#,##0.00",None,None]
        for j, (v, fmt) in enumerate(zip(vals, fmts)):
            c = ws2.cell(r, j+1, v)
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center" if j in (0,2,5,6,7,8) else "left", vertical="center")
            if fmt: c.number_format = fmt
        ws2.row_dimensions[r].height = 13

    # Totals
    rt = len(assets) + 2
    ws2.cell(rt, 1, "TOTALS").font = Font(name="Arial", bold=True, color=WHITE, size=9)
    ws2.cell(rt, 1).fill = PatternFill("solid", fgColor=DARK_BLUE)
    for col, val, fmt in [
        (5, sum(float(a["COST"] or 0) for a in active), "£#,##0.00"),
        (6, sum(dep5(a["PURCHASE_DATE"], float(a["COST"] or 0), TODAY.year, TODAY.month) for a in active), "£#,##0.00"),
        (7, sum(float(a["ACCOUNTING_BOOK_VALUE"] or 0) for a in active), "£#,##0.00"),
    ]:
        c = ws2.cell(rt, col, val)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.number_format = fmt
        c.alignment = Alignment(horizontal="center")

    # ────────────────────────────────────────────────
    # TAB 3: MONTHLY DEP SCHEDULE (all history)
    # ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Monthly Dep Schedule")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 5
    for i in range(len(all_months)):
        ws3.column_dimensions[get_column_letter(i + 4)].width = 7
    ws3.column_dimensions[get_column_letter(len(all_months) + 4)].width = 10

    # Title
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_months)+4)
    c = ws3.cell(1, 1, f"RETRO LABS LTD — Monthly Depreciation — Oct 2021 to Jul 2026 — Auto-generated {TODAY}")
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 18

    # Year banners (row 2) + month labels (row 3)
    for col, label in [(1,""),(2,""),(3,"")]:
        hdr(ws3, 2, col, label, DARK_BLUE, WHITE, size=8)
        hdr(ws3, 3, col, ["FA #","Asset Name","Type"][col-1], DARK_BLUE, WHITE, size=8)
    hdr(ws3, 2, len(all_months)+4, "", DARK_BLUE, WHITE, size=8)
    hdr(ws3, 3, len(all_months)+4, "Total £", MID_BLUE, WHITE, size=8)

    yr_start = None; yr_start_col = None
    for i, (yr, mo) in enumerate(all_months):
        col = i + 4
        if yr != yr_start:
            if yr_start is not None:
                ws3.merge_cells(start_row=2, start_column=yr_start_col, end_row=2, end_column=col-1)
                c = ws3.cell(2, yr_start_col, str(yr_start))
                c.font = Font(name="Arial", bold=True, color=WHITE, size=8)
                c.fill = PatternFill("solid", fgColor=MID_BLUE if yr_start % 2 == 0 else DARK_BLUE)
                c.alignment = Alignment(horizontal="center", vertical="center")
            yr_start = yr; yr_start_col = col
        mo_bg = "375623" if (yr, mo) in dep_journals else ("2E75B6" if yr >= 2025 else "4472C4")
        c = ws3.cell(3, col, datetime.date(yr, mo, 1).strftime("%b"))
        c.font = Font(name="Arial", bold=True, color=WHITE, size=7)
        c.fill = PatternFill("solid", fgColor=mo_bg)
        c.alignment = Alignment(horizontal="center")
    # Last year banner
    ws3.merge_cells(start_row=2, start_column=yr_start_col, end_row=2, end_column=len(all_months)+3)
    c = ws3.cell(2, yr_start_col, str(yr_start))
    c.font = Font(name="Arial", bold=True, color=WHITE, size=8)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center")
    ws3.row_dimensions[2].height = 14; ws3.row_dimensions[3].height = 13

    for i, a in enumerate(assets):
        r = i + 4
        bg = RED_LIGHT if a["ASSET_STATUS"] == "Disposed" else (LIGHT_BLUE if i % 2 == 0 else WHITE)
        ws3.cell(r, 1, a["ASSET_NUMBER"]).fill = PatternFill("solid", fgColor=bg)
        ws3.cell(r, 1).font = Font(name="Arial", size=7)
        ws3.cell(r, 1).alignment = Alignment(horizontal="center")
        ws3.cell(r, 2, a["ASSET_NAME"]).fill = PatternFill("solid", fgColor=bg)
        ws3.cell(r, 2).font = Font(name="Arial", size=7)
        ws3.cell(r, 3, "CE" if "Computer" in (a.get("ASSET_TYPE_NAME") or "") else "OE")
        ws3.cell(r, 3).fill = PatternFill("solid", fgColor=bg)
        ws3.cell(r, 3).font = Font(name="Arial", size=7)
        ws3.cell(r, 3).alignment = Alignment(horizontal="center")

        row_total = 0.0
        pd = a["PURCHASE_DATE"]; cost = float(a["COST"] or 0)
        for j, (yr, mo) in enumerate(all_months):
            col = j + 4
            d = dep5(pd, cost, yr, mo, a.get("DISPOSAL_DATE"))
            cell_bg = "C6EFCE" if (yr,mo) in dep_journals and d > 0 else bg
            c = ws3.cell(r, col, d if d > 0 else None)
            c.font = Font(name="Arial", size=7)
            c.fill = PatternFill("solid", fgColor=cell_bg)
            c.alignment = Alignment(horizontal="center")
            if d > 0: c.number_format = "£#,##0.00"
            row_total += d
        c = ws3.cell(r, len(all_months)+4, row_total if row_total > 0 else None)
        c.font = Font(name="Arial", size=7, bold=True)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center")
        if row_total > 0: c.number_format = "£#,##0.00"
        ws3.row_dimensions[r].height = 11

    # Totals row
    rt3 = len(assets) + 4
    ws3.merge_cells(start_row=rt3, start_column=1, end_row=rt3, end_column=3)
    ws3.cell(rt3, 1, "MONTHLY TOTAL").font = Font(name="Arial", bold=True, color=WHITE, size=8)
    ws3.cell(rt3, 1).fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws3.cell(rt3, 1).alignment = Alignment(horizontal="center")
    for j, (yr, mo) in enumerate(all_months):
        col = j + 4
        total = sum(dep5(a["PURCHASE_DATE"], float(a["COST"] or 0), yr, mo, a.get("DISPOSAL_DATE")) for a in assets)
        c = ws3.cell(rt3, col, total)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=7)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.number_format = "£#,##0.00"
        c.alignment = Alignment(horizontal="center")
    ws3.row_dimensions[rt3].height = 14

    # ────────────────────────────────────────────────
    # TAB 4: XERO RECONCILIATION
    # ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Xero Reconciliation")
    ws4.sheet_view.showGridLines = False
    for col, w in [(1,5),(2,22),(3,14),(4,14),(5,14),(6,18)]:
        ws4.column_dimensions[get_column_letter(col)].width = w

    ws4.merge_cells("A1:F1")
    ws4["A1"].value = f"XERO ACTUAL vs SCHEDULE — Last {min(6, len(dep_journals))} periods — Updated {TODAY}"
    ws4["A1"].font = Font(name="Arial", bold=True, size=10, color=WHITE)
    ws4["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 20

    for col, title in enumerate(["Period","Xero CE £","Xero OE £","Xero Total £","Schedule £","Variance £"], 1):
        hdr(ws4, 2, col, title, MID_BLUE, WHITE, size=9)
    ws4.row_dimensions[2].height = 16

    for i, key in enumerate(sorted(dep_journals.keys(), reverse=True)[:12]):
        r = i + 3
        d = dep_journals[key]
        period_str = datetime.date(key[0], key[1], 1).strftime("%b %Y")
        sched = sum(dep5(a["PURCHASE_DATE"], float(a["COST"] or 0), key[0], key[1],
                         a.get("DISPOSAL_DATE")) for a in assets)
        variance = abs(d["total"] - sched)
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        s_bg = GREEN if variance < 10 else AMBER
        dat(ws4, r, 1, period_str, bg=bg, bold=True, size=9, center=True)
        dat(ws4, r, 2, d["CE"],    bg=bg, size=9, center=True, fmt="£#,##0.00")
        dat(ws4, r, 3, d["OE"],    bg=bg, size=9, center=True, fmt="£#,##0.00")
        dat(ws4, r, 4, d["total"], bg=bg, bold=True, size=9, center=True, fmt="£#,##0.00")
        dat(ws4, r, 5, sched,      bg=bg, size=9, center=True, fmt="£#,##0.00")
        dat(ws4, r, 6, variance,   bg=s_bg, bold=True, size=9, center=True, fmt="£#,##0.00")
        ws4.row_dimensions[r].height = 14

    return wb


# ─── GOOGLE DRIVE UPLOAD ──────────────────────────────────────────────────────

def upload_to_drive(wb) -> str:
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    folder_id = os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "")
    if not sa_json or not folder_id:
        print("  [Drive] No credentials — skipping upload.")
        return ""

    creds = service_account.Credentials.from_service_account_info(
        json.loads(sa_json),
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    service = gdrive_build("drive", "v3", credentials=creds)

    LIVE_NAME    = "FA_Schedule_FINAL.xlsx"
    MIMETYPE     = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # ── Find the current live file (if it exists) ──────────────────────────
    results = service.files().list(
        q=f"name='{LIVE_NAME}' and '{folder_id}' in parents and trashed=false",
        fields="files(id,name)"
    ).execute()
    existing = results.get("files", [])

    if existing:
        live_id = existing[0]["id"]

        # ── Archive the current version with last month's label ────────────
        # e.g. FA_Schedule_2026_02.xlsx  (the file we are about to replace
        # contains last month's data, so label it with the prior period)
        prior = TODAY - relativedelta_simple(datetime.date(TODAY.year, TODAY.month, 1))
        # relativedelta_simple advances; we need to go back one month instead
        if TODAY.month == 1:
            archive_yr, archive_mo = TODAY.year - 1, 12
        else:
            archive_yr, archive_mo = TODAY.year, TODAY.month - 1
        archive_name = f"FA_Schedule_{archive_yr}_{archive_mo:02d}.xlsx"

        # Only archive if that month's copy doesn't already exist
        existing_archive = service.files().list(
            q=f"name='{archive_name}' and '{folder_id}' in parents and trashed=false",
            fields="files(id)"
        ).execute().get("files", [])

        if not existing_archive:
            # Copy the current live file and rename it as the archive
            copy_meta = {"name": archive_name, "parents": [folder_id]}
            copied = service.files().copy(fileId=live_id, body=copy_meta, fields="id").execute()
            print(f"  [Drive] Archived prior version as '{archive_name}' (id: {copied['id']})")
        else:
            print(f"  [Drive] Archive '{archive_name}' already exists — skipping.")

        # ── Overwrite the live file with the new version ───────────────────
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        media = MediaIoBaseUpload(buf, mimetype=MIMETYPE, resumable=True)
        service.files().update(fileId=live_id, media_body=media).execute()
        print(f"  [Drive] Updated live file '{LIVE_NAME}' (id: {live_id})")
        file_id = live_id

    else:
        # ── First-ever upload — create the live file ───────────────────────
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        media = MediaIoBaseUpload(buf, mimetype=MIMETYPE, resumable=True)
        meta  = {"name": LIVE_NAME, "parents": [folder_id]}
        f     = service.files().create(body=meta, media_body=media, fields="id").execute()
        file_id = f["id"]
        print(f"  [Drive] Created live file '{LIVE_NAME}' (id: {file_id})")

    return f"https://drive.google.com/file/d/{file_id}/view"


# ─── SIMPLE RELATIVEDELTA ─────────────────────────────────────────────────────

def relativedelta_simple(d):
    """Advance a date by one month."""
    if d.month == 12:
        return datetime.date(d.year + 1, 1, 1)
    return datetime.date(d.year, d.month + 1, 1)


# ─── ENTRY POINT ─────────────────────────────────────────────────────────────

def run(conn=None) -> str:
    """Called from fa_depreciation.py step 5. Returns Google Drive URL."""
    close_conn = False
    if conn is None:
        from fa_depreciation import get_conn
        conn = get_conn()
        close_conn = True
    try:
        print("  Fetching assets from Snowflake...")
        assets = fetch_assets(conn)
        print(f"  {len(assets)} assets fetched.")
        print("  Fetching posted dep journals...")
        dep_journals = fetch_dep_journals(conn)
        print(f"  {len(dep_journals)} periods of journal data fetched.")
        print("  Building workbook...")
        wb = build_workbook(assets, dep_journals)
        print("  Uploading to Google Drive...")
        url = upload_to_drive(wb)
        return url
    finally:
        if close_conn:
            conn.close()


if __name__ == "__main__":
    # Standalone run — useful for testing
    from dotenv import load_dotenv
    load_dotenv()
    url = run()
    print(f"\nDone. Drive URL: {url}")
